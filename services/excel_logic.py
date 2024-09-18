from datetime import timedelta
from io import BytesIO

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def create_events_excel(events: list):
    book = Workbook()
    sheet: Worksheet = book.active

    sheet.append(
        [
            "ID",
            "Название мероприятия",
            "Описание",
            "Дата и время начала",
            "Дата и время окончания",
            "Тэги",
            "Адрес",
            "Организация",
            "Эко-баллы",
            "Количество участников",
        ]
    )
    for event in events:
        sheet.append(
            [
                event.id,
                event.name,
                event.description,
                event.start_date.replace(tzinfo=None) + timedelta(hours=3),
                event.finish_date.replace(tzinfo=None) + timedelta(hours=3),
                ";".join([tag.name for tag in event.tags.all()]),
                event.address,
                event.organization.full_name,
                event.eco_balance,
                event.tickets.count(),
            ]
        )

    file = BytesIO()
    book.save(file)
    file.seek(0)
    return file


def create_users_excel(users: list):
    book = Workbook()
    sheet: Worksheet = book.active

    sheet.append(["Почта", "ФИО", "Эко-баллы", "Количество посещенных мероприятий"])
    for user in users:
        sheet.append([user.email, user.full_name, user.eco_balance, user.tickets.count()])

    file = BytesIO()
    book.save(file)
    file.seek(0)
    return file


def create_organizations_excel(organizations: list):
    book = Workbook()
    sheet: Worksheet = book.active

    sheet.append(
        [
            "Полное название",
            "Короткое название",
            "Адрес",
            "ОРГН",
            "ИНН",
            "КПП",
            "ОКПО",
            "Директор ФИО",
            "Директор ИНН",
            "Виды деятельности",
            "Дата регистрации",
            "Телефон",
            "Email",
            "Описание",
            "Число людей в организации",
            "Число мероприятий",
        ]
    )
    for organization in organizations:
        sheet.append(
            [
                organization.full_name,
                organization.short_name,
                organization.address,
                organization.ogrn,
                organization.inn,
                organization.kpp,
                organization.okpo,
                organization.director_full_name,
                organization.director_inn,
                organization.activities,
                organization.reg_date,
                organization.phone,
                organization.email,
                organization.description,
                organization.users.count(),
                organization.events.count(),
            ]
        )

    file = BytesIO()
    book.save(file)
    file.seek(0)
    return file
